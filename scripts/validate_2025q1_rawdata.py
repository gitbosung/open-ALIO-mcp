# -*- coding: utf-8 -*-
"""Validate current parsed ALIO data against 2025 Q1 raw XLSX files.

The validation folder supplied by the user contains 2025 Q1 item-level exports.
The current crawl in this repository is 2026 Q1, so numeric checks intentionally
compare only the closed overlapping years 2021-2024. Newest-quarter columns are
reported as coverage context, not used as strict equality evidence.
"""
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
VALIDATION_DIR = ROOT / "rawdata" / "validation"
HTML_DIR = ROOT / "rawdata" / "html"
CRAWL_CSV = ROOT / "data" / "crawl" / "alio_records.csv"
METRICS_DIR = ROOT / "data" / "metrics"
INSTITUTIONS = ROOT / "data" / "institutions.json"
REPORT_DIR = ROOT / "data" / "validation_reports"

COMPARE_YEARS = {"2021", "2022", "2023", "2024"}
META_COLS = {"기관명", "기관유형", "주무부처", "상위기관"}
YEAR_RE = re.compile(r"(20\d{2})")


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("&cr;", " ")).strip()


def compact(value: object) -> str:
    return re.sub(r"\s+", "", clean(value))


def to_num(value: object) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        return int(f) if f.is_integer() else round(f, 6)
    text = clean(value).replace(",", "").replace("%", "")
    if text in {"", "-", "–", "—", "해당사항 없음"}:
        return None
    try:
        f = float(text)
    except ValueError:
        return None
    return int(f) if f.is_integer() else round(f, 6)


def approx_equal(left: object, right: object) -> bool:
    try:
        return abs(float(left) - float(right)) <= 0.000001
    except (TypeError, ValueError):
        return clean(left) == clean(right)


def load_name_to_code() -> dict[str, str]:
    data = json.loads(INSTITUTIONS.read_text(encoding="utf-8"))
    return {org["name"]: org["org_code"] for org in data["orgs"] if org.get("org_code")}


def find_file(token: str) -> Path:
    matches = sorted(path for path in VALIDATION_DIR.rglob("*.xlsx") if token in path.name)
    if not matches:
        raise FileNotFoundError(token)
    return matches[0]


def header_row(ws) -> tuple[int, list[str]] | None:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
        values = [clean(c.value) for c in row]
        if "기관명" in values:
            return row[0].row, values
    return None


def iter_table_rows(path: Path) -> Iterable[dict[str, object]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            hdr = header_row(ws)
            if not hdr:
                continue
            hdr_row, headers = hdr
            for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
                rec = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
                if clean(rec.get("기관명")):
                    yield rec
    finally:
        wb.close()


def year_columns(row: dict[str, object]) -> list[tuple[str, str]]:
    cols: list[tuple[str, str]] = []
    for header in row:
        if header in META_COLS:
            continue
        match = YEAR_RE.search(header)
        if match:
            cols.append((header, match.group(1)))
    return cols


def extract_wide(
    file_token: str,
    category: str,
    key_fn: Callable[[dict[str, object]], str | None],
    item_no: str | None = None,
) -> dict:
    path = find_file(file_token)
    records = []
    for row in iter_table_rows(path):
        org_name = clean(row.get("기관명"))
        key = key_fn(row)
        if not key:
            continue
        for header, year in year_columns(row):
            value = to_num(row.get(header))
            if value is None:
                continue
            records.append({
                "file": path.name,
                "category": category,
                "org_name": org_name,
                "metric_key": key,
                "year": year,
                "value": value,
            })
    return {"file": path.name, "category": category, "item_no": item_no, "records": records}


def extract_tax() -> dict:
    path = find_file("37-1")
    cols = ["과세표준", "법인세 산출세액", "세액공제", "가산세", "결정세액"]
    records = []
    for row in iter_table_rows(path):
        year = YEAR_RE.search(clean(row.get("연도") or row.get("년도")))
        if not year or year.group(1) not in COMPARE_YEARS:
            continue
        for col in cols:
            value = to_num(row.get(col))
            if value is None:
                continue
            records.append({
                "file": path.name,
                "category": "tax",
                "org_name": clean(row.get("기관명")),
                "metric_key": col,
                "year": year.group(1),
                "value": value,
            })
    return {"file": path.name, "category": "tax", "item_no": "32211", "records": records}


def extract_welfare_etc_leave_pay() -> dict:
    path = find_file("14-1-2")
    records = []
    for row in iter_table_rows(path):
        year = YEAR_RE.search(clean(row.get("년도") or row.get("연도")))
        value = to_num(row.get("지급금액"))
        if not year or value is None:
            continue
        key = f"휴직급여지급현황 | {clean(row.get('구분'))} | {clean(row.get('사유'))}"
        records.append({
            "file": path.name,
            "category": "welfare_etc",
            "org_name": clean(row.get("기관명")),
            "metric_key": key,
            "year": year.group(1),
            "value": value,
        })
    return {"file": path.name, "category": "welfare_etc", "item_no": "63701", "records": records}


def budget_key(row: dict[str, object], fund: bool) -> str | None:
    label = clean(row.get("항목"))
    if not label:
        return None
    kind = "정부순지원수입" if label.startswith("정부순지원수입") else "수입지출현황"
    parts = [f"{kind}({'기금계정' if fund else '고유사업'})"]
    if fund:
        account = clean(row.get("기금명"))
        if account:
            parts.append(account)
    parts.append(label)
    return " | ".join(parts)


def exact_sources() -> list[dict]:
    sources = [
        extract_wide("3-1. 임직원수(", "staff", lambda r: clean(r.get("분류")), "20201"),
        extract_wide("11-1", "salary", lambda r: f"직원평균보수 | {clean(r.get('분류1'))} | {clean(r.get('분류2'))}", "20601"),
        extract_wide("11-2", "salary", lambda r: f"신입사원초임 | {clean(r.get('분류'))}", "20601"),
        extract_wide("10.", "executive_pay", lambda r: f"{clean(r.get('구분'))} | {clean(r.get('분류'))}", "20501"),
        extract_wide("12.", "head_expense", lambda r: clean(r.get("항목")), "20701"),
        extract_wide("5-1. 신규채용현황", "recruitment", lambda r: f"신규채용현황 | {compact(r.get('분류'))}", "20401"),
        extract_wide("5-1. 총신규채용", "recruitment", lambda r: f"신규채용현황 | {compact(r.get('분류'))}", "20401"),
        extract_wide("5-2", "recruitment", lambda r: f"청년인턴채용현황 | {compact(r.get('분류'))}", "20401"),
        extract_wide("13-1", "welfare", lambda r: f"예산상 복리후생비 | {clean(r.get('고용형태'))} | {clean(r.get('항목'))}", "20801"),
        extract_wide("24-1", "work_life", lambda r: f"일가정-육아휴직사용자수 | {clean(r.get('분류'))}", "21401"),
        extract_wide("24-2", "work_life", lambda r: f"일가정-출산휴가사용자수 | {clean(r.get('분류'))}", "21401"),
        extract_wide("24-3", "work_life", lambda r: f"일가정-임신기육아기단축근무제사용자수 | {clean(r.get('분류'))}", "21401"),
        extract_wide("24-5", "work_life", lambda r: f"일가정-가족돌봄휴가사용자수 | {clean(r.get('구분'))}", "21401"),
        extract_wide("24-6", "work_life", lambda r: f"일가정-가족돌봄휴직사용자수 | {clean(r.get('구분'))}", "21401"),
        extract_wide("33. 수입지출현황[고유사업]", "budget", lambda r: budget_key(r, False), "31401"),
        extract_wide("33. 수입지출현황[기금계정]", "budget", lambda r: budget_key(r, True), "31401"),
        extract_tax(),
        extract_welfare_etc_leave_pay(),
    ]
    return sources


def canonical_key(value: str) -> str:
    text = compact(value)
    text = re.sub(r"\([A-Z/]+\)", "", text)
    return (
        text.replace("급여성복리후생비", "급여성")
        .replace("비급여성복리후생비", "비급여성")
        .replace("정규직(일반)", "정규직(일반정규직)")
        .replace("정규직(무기)", "정규직(무기계약직)")
        .replace("법인세산출세액", "법인세산출세액")
    )


def row_label_flat(row_label: str) -> str:
    return clean(row_label).replace(" > ", "-").replace(" ", "")


def salary_row_label(row_label: str) -> str:
    label = clean(row_label)
    label = label.replace(" > (남성)", " - 남성").replace(" > (여성)", " - 여성")
    label = label.replace("평균 보수액", "평균보수액").replace("평균 근속연수", "평균근속연수")
    return label


def csv_compare_key(row: dict[str, str]) -> str | None:
    item = row["item_no"]
    label = clean(row.get("row_label"))
    col = clean(row.get("col_label"))
    section = re.sub(r"^\d+[-.)]?\s*", "", clean(row.get("section")))
    if item == "20201":
        return row_label_flat(label)
    if item == "20601":
        if "신입" in section:
            return f"신입사원초임 | {salary_row_label(label)}"
        return f"직원평균보수 | {section} | {salary_row_label(label)}"
    if item == "20501":
        return f"{section} | {label}"
    if item == "20701":
        return col if col == "업무추진비 집행금액" else None
    if item == "20401":
        prefix = "청년인턴채용현황" if "청년인턴" in section else "신규채용현황"
        return f"{prefix} | {row_label_flat(label)}"
    if item == "20801":
        return f"예산상 복리후생비 | {section} | {label}"
    if item == "21401":
        section_key = {
            "육아휴직 사용자 수 및 사용률": "일가정-육아휴직사용자수",
            "출산 휴가 사용자 수 및 사용률": "일가정-출산휴가사용자수",
            "임신기·육아기 단축 근무제 사용자 수": "일가정-임신기육아기단축근무제사용자수",
            "가족돌봄휴가 사용자수": "일가정-가족돌봄휴가사용자수",
            "가족돌봄휴직 사용자수": "일가정-가족돌봄휴직사용자수",
        }.get(section)
        return f"{section_key} | {label}" if section_key else None
    if item == "31401":
        sub_account = clean(row.get("sub_account"))
        kind = "정부순지원수입" if label.startswith("정부순지원수입") else "수입지출현황"
        account_type = "기금계정" if sub_account or "기금" in section else "고유사업"
        parts = [f"{kind}({account_type})"]
        if sub_account:
            parts.append(sub_account)
        parts.append(label)
        return " | ".join(parts)
    if item == "32211":
        return col.replace("법인세산출세액", "법인세 산출세액")
    return None


def compare_to_metrics(source: dict, name_to_code: dict[str, str]) -> dict:
    metric = json.loads((METRICS_DIR / f"{source['category']}.json").read_text(encoding="utf-8"))
    orgs = metric.get("orgs", {})
    stats = Counter()
    samples = []
    years_seen = set()
    keys_seen = set()
    orgs_seen = set()

    for rec in source["records"]:
        year = str(rec["year"])
        years_seen.add(year)
        keys_seen.add(rec["metric_key"])
        if year not in COMPARE_YEARS:
            stats["out_of_period"] += 1
            continue
        code = name_to_code.get(rec["org_name"])
        if not code:
            stats["unmatched_org"] += 1
            if len(samples) < 10:
                samples.append({**rec, "reason": "unmatched_org"})
            continue
        orgs_seen.add(code)
        series = orgs.get(code, {}).get("series", {})
        values = series.get(rec["metric_key"])
        if values is None:
            stats["missing_key"] += 1
            if len(samples) < 10:
                samples.append({**rec, "org_code": code, "reason": "missing_key"})
            continue
        actual = values.get(year)
        if actual is None:
            stats["missing_year"] += 1
            if len(samples) < 10:
                samples.append({**rec, "org_code": code, "reason": "missing_year"})
            continue
        if approx_equal(actual, rec["value"]):
            stats["matched"] += 1
        else:
            stats["mismatched"] += 1
            if len(samples) < 10:
                samples.append({**rec, "org_code": code, "actual": actual, "reason": "mismatched"})

    strict_total = stats["matched"] + stats["mismatched"] + stats["missing_key"] + stats["missing_year"]
    return {
        "file": source["file"],
        "category": source["category"],
        "records": len(source["records"]),
        "compare_years": sorted(COMPARE_YEARS),
        "years_seen": sorted(years_seen),
        "org_count": len(orgs_seen),
        "key_count": len(keys_seen),
        "matched": stats["matched"],
        "mismatched": stats["mismatched"],
        "missing_key": stats["missing_key"],
        "missing_year": stats["missing_year"],
        "unmatched_org": stats["unmatched_org"],
        "out_of_period": stats["out_of_period"],
        "strict_total": strict_total,
        "match_rate": round(stats["matched"] / strict_total, 6) if strict_total else None,
        "samples": samples,
    }


def build_csv_index(item_nos: set[str]) -> dict[tuple[str, str, str], set[int | float | None]]:
    index: dict[tuple[str, str, str], set[int | float | None]] = defaultdict(set)
    with CRAWL_CSV.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            if row["item_no"] not in item_nos or row["year"] not in COMPARE_YEARS:
                continue
            key = csv_compare_key(row)
            if not key:
                continue
            value = to_num(row.get("value"))
            index[(row["apba_id"], row["year"], canonical_key(key))].add(value)
    return index


def compare_to_crawl_csv(sources: list[dict], name_to_code: dict[str, str]) -> list[dict]:
    item_nos = {src["item_no"] for src in sources if src.get("item_no")}
    index = build_csv_index(item_nos)
    results = []
    for source in sources:
        if not source.get("item_no"):
            continue
        stats = Counter()
        samples = []
        orgs_seen = set()
        keys_seen = set()
        for rec in source["records"]:
            year = str(rec["year"])
            if year not in COMPARE_YEARS:
                stats["out_of_period"] += 1
                continue
            code = name_to_code.get(rec["org_name"])
            if not code:
                stats["unmatched_org"] += 1
                continue
            orgs_seen.add(code)
            ckey = canonical_key(rec["metric_key"])
            keys_seen.add(ckey)
            values = index.get((code, year, ckey))
            if not values:
                if rec["value"] == 0:
                    stats["zero_missing_csv_key"] += 1
                    continue
                stats["missing_csv_key"] += 1
                if len(samples) < 10:
                    samples.append({**rec, "org_code": code, "reason": "missing_csv_key"})
                continue
            if rec["value"] == 0 and None in values:
                stats["zero_blank_tolerated"] += 1
                continue
            if any(approx_equal(value, rec["value"]) for value in values):
                stats["matched"] += 1
            else:
                stats["mismatched"] += 1
                if len(samples) < 10:
                    sortable = sorted("" if v is None else v for v in values)
                    samples.append({**rec, "org_code": code, "csv_values": sortable, "reason": "mismatched"})
        strict_total = stats["matched"] + stats["mismatched"] + stats["missing_csv_key"]
        results.append({
            "file": source["file"],
            "item_no": source.get("item_no"),
            "category": source["category"],
            "records": len(source["records"]),
            "org_count": len(orgs_seen),
            "key_count": len(keys_seen),
            "matched": stats["matched"],
            "mismatched": stats["mismatched"],
            "missing_csv_key": stats["missing_csv_key"],
            "zero_blank_tolerated": stats["zero_blank_tolerated"],
            "zero_missing_csv_key": stats["zero_missing_csv_key"],
            "unmatched_org": stats["unmatched_org"],
            "out_of_period": stats["out_of_period"],
            "strict_total": strict_total,
            "match_rate": round(stats["matched"] / strict_total, 6) if strict_total else None,
            "samples": samples,
        })
    return results


def workbook_inventory() -> dict:
    files = sorted(VALIDATION_DIR.rglob("*.xlsx"))
    broken = []
    sheet_count = 0
    main_sheet_count = 0
    rows_with_org_header = 0
    for path in files:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - diagnostic script
            broken.append({"file": path.name, "error": str(exc)})
            continue
        try:
            sheet_count += len(wb.worksheets)
            for ws in wb.worksheets:
                hdr = header_row(ws)
                if hdr:
                    main_sheet_count += 1
                    rows_with_org_header += max(0, ws.max_row - hdr[0])
        finally:
            wb.close()
    return {
        "xlsx_count": len(files),
        "sheet_count": sheet_count,
        "main_sheet_count": main_sheet_count,
        "rows_with_org_header": rows_with_org_header,
        "broken": broken,
    }


def crawl_summary() -> dict:
    doc_counts = Counter(path.name.split("_")[1] for path in HTML_DIR.glob("*__doc.html"))
    csv_counts = Counter()
    csv_orgs: dict[str, set[str]] = defaultdict(set)
    csv_as_of = Counter()
    with CRAWL_CSV.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            item = row["item_no"]
            csv_counts[item] += 1
            csv_orgs[item].add(row["apba_id"])
            if row.get("as_of"):
                csv_as_of[row["as_of"]] += 1
    items = []
    for item in sorted(set(doc_counts) | set(csv_counts)):
        items.append({
            "item_no": item,
            "doc_count": doc_counts[item],
            "csv_rows": csv_counts[item],
            "csv_org_count": len(csv_orgs[item]),
        })
    return {
        "as_of_counts": csv_as_of.most_common(),
        "total_csv_rows": sum(csv_counts.values()),
        "item_count_with_docs": len(doc_counts),
        "item_count_with_csv": len(csv_counts),
        "items": items,
        "doc_items_with_zero_csv": [i for i in items if i["doc_count"] and not i["csv_rows"]],
    }


VALIDATION_FILE_ITEM_MAP = {
    "10. 임원연봉현황": "20501",
    "11-1. 직원평균보수": "20601",
    "11-2. 신입사원초임": "20601",
    "12. 기관장업무추진비": "20701",
    "13-1. 예산상복리후생비": "20801",
    "13-2. 사내복지기금": "20801",
    "13-3": "20801",
    "13-5. 1인당복리후생비": "20801",
    "14-": "63701",
    "15. 노동조합가입정보": "21021",
    "16-1": "21301",
    "16-2": "21311",
    "19-1. 에너지": "21621",
    "19-2. 폐기물": "21631",
    "19-3. 용수": "21641",
    "21-1. 산업재해": "70461",
    "24-": "21401",
    "29. 자체 감사부서": "32311",
    "3-1. 임직원": "20201",
    "3-2. 직급별": "20201",
    "30. 청렴도": "40211",
    "31. 요약재무상태표": "31201",
    "32. 요약손익계산서": "31301",
    "33. 수입지출현황": "31401",
    "34. 자본금": "31701",
    "35. 장단기": "31801",
    "36-1. 외부회계": "32301",
    "37-1. 법인세": "32211",
    "38. 주요사업": "31501",
    "39. 투자집행": "31601",
    "40-1": "31901",
    "40-3": "31901",
    "40-5": "32001",
    "41-": "32101",
    "47-2. 수의계약": "70301",
    "5-1": "20401",
    "5-2": "20401",
    "7-1": "31921",
    "7-2": "21801",
    "8-1": "21201",
    "8-2": "21211",
}


def mapped_item_for_file(name: str) -> str | None:
    for token, item in sorted(VALIDATION_FILE_ITEM_MAP.items(), key=lambda x: len(x[0]), reverse=True):
        if name.startswith(token):
            return item
    return None


def validation_coverage(crawl: dict) -> dict:
    doc_by_item = {item["item_no"]: item["doc_count"] for item in crawl["items"]}
    csv_by_item = {item["item_no"]: item["csv_rows"] for item in crawl["items"]}
    files = []
    for path in sorted(VALIDATION_DIR.rglob("*.xlsx")):
        item = mapped_item_for_file(path.name)
        files.append({
            "file": path.name,
            "mapped_item_no": item,
            "has_html_docs": bool(item and doc_by_item.get(item)),
            "html_doc_count": doc_by_item.get(item, 0) if item else 0,
            "csv_rows": csv_by_item.get(item, 0) if item else 0,
        })
    return {
        "files_total": len(files),
        "mapped_files": sum(1 for f in files if f["mapped_item_no"]),
        "mapped_to_crawled_item": sum(1 for f in files if f["has_html_docs"]),
        "mapped_to_parsed_csv": sum(1 for f in files if f["csv_rows"]),
        "files": files,
    }


def write_reports(report: dict) -> tuple[Path, Path]:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = REPORT_DIR / "2025q1_rawdata_validation.json"
    md_path = REPORT_DIR / "2025q1_rawdata_validation.md"
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# 2025 Q1 Rawdata Validation",
        "",
        f"- generated_at: {report['generated_at']}",
        f"- validation_xlsx_count: {report['inventory']['xlsx_count']}",
        f"- crawl_as_of: {report['crawl']['as_of_counts'][:3]}",
        f"- crawl_csv_rows: {report['crawl']['total_csv_rows']:,}",
        f"- doc_items_with_zero_csv: {report['crawl']['doc_items_with_zero_csv']}",
        "",
        "## Exact Metric Checks",
        "",
        "| file | category | matched | mismatched | missing_key | missing_year | match_rate |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for item in report["metric_checks"]:
        rate = "" if item["match_rate"] is None else f"{item['match_rate']:.4%}"
        lines.append(
            f"| {item['file']} | {item['category']} | {item['matched']:,} | "
            f"{item['mismatched']:,} | {item['missing_key']:,} | {item['missing_year']:,} | {rate} |"
        )
    lines.extend([
        "",
        "## Direct Crawl CSV Checks",
        "",
        "| file | item_no | matched | mismatched | missing_csv_key | zero_blank_tolerated | match_rate |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: |",
    ])
    for item in report["crawl_csv_checks"]:
        rate = "" if item["match_rate"] is None else f"{item['match_rate']:.4%}"
        lines.append(
            f"| {item['file']} | {item['item_no']} | {item['matched']:,} | "
            f"{item['mismatched']:,} | {item['missing_csv_key']:,} | "
            f"{item['zero_blank_tolerated'] + item['zero_missing_csv_key']:,} | {rate} |"
        )
    lines.extend([
        "",
        "## Notes",
        "",
        "- Strict equality is limited to 2021-2024 because current parsed HTML is 2026 Q1 while the rawdata folder is 2025 Q1.",
        "- Direct crawl CSV checks compare the supplied XLSX records against `data/crawl/alio_records.csv`, not the derived metric JSON.",
        "- 2020 and newest-quarter columns are inventory context only.",
        "- Missing keys usually mean the current MCP metric layer has not promoted that raw table shape yet.",
    ])
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return json_path, md_path


def main() -> int:
    name_to_code = load_name_to_code()
    inventory = workbook_inventory()
    crawl = crawl_summary()
    coverage = validation_coverage(crawl)
    sources = exact_sources()
    metric_checks = [compare_to_metrics(src, name_to_code) for src in sources]
    crawl_csv_checks = compare_to_crawl_csv(sources, name_to_code)

    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "scope": {
            "validation_dir": str(VALIDATION_DIR.relative_to(ROOT)),
            "validation_period": "2025 Q1 rawdata",
            "current_crawl_period": "2026 Q1 parsed HTML",
            "strict_compare_years": sorted(COMPARE_YEARS),
        },
        "inventory": inventory,
        "crawl": crawl,
        "validation_coverage": coverage,
        "crawl_csv_checks": crawl_csv_checks,
        "metric_checks": metric_checks,
    }
    json_path, md_path = write_reports(report)

    print(f"validation files: {inventory['xlsx_count']} xlsx, broken={len(inventory['broken'])}")
    print(f"crawl csv rows: {crawl['total_csv_rows']:,}; items with docs={crawl['item_count_with_docs']}; items with csv={crawl['item_count_with_csv']}")
    print(f"doc items with zero parsed csv: {crawl['doc_items_with_zero_csv']}")
    print("exact metric checks:")
    for item in metric_checks:
        rate = "n/a" if item["match_rate"] is None else f"{item['match_rate']:.4%}"
        print(
            f"  {item['file']} -> {item['category']}: matched={item['matched']:,} "
            f"mismatch={item['mismatched']:,} missing_key={item['missing_key']:,} "
            f"missing_year={item['missing_year']:,} rate={rate}"
        )
        for sample in item["samples"][:2]:
            print(f"    sample {sample}")
    print("direct crawl csv checks:")
    for item in crawl_csv_checks:
        rate = "n/a" if item["match_rate"] is None else f"{item['match_rate']:.4%}"
        print(
            f"  {item['file']} -> item {item['item_no']}: matched={item['matched']:,} "
            f"mismatch={item['mismatched']:,} missing_csv_key={item['missing_csv_key']:,} "
            f"zero_tolerated={item['zero_blank_tolerated'] + item['zero_missing_csv_key']:,} "
            f"rate={rate}"
        )
        for sample in item["samples"][:2]:
            print(f"    sample {sample}")
    print(f"json report: {json_path.relative_to(ROOT)}")
    print(f"md report: {md_path.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
