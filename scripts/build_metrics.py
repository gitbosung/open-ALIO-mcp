# -*- coding: utf-8 -*-
"""rawdata/xlsx/ 의 ALIO 항목별 공시 엑셀 → data/metrics/*.json

공통 구조 (조회 최적화):
{
  "_meta": {category, label, unit, source_files, years, built_at, caveats},
  "orgs": {
    "<org_code>": {
      "name": "한국전력공사",
      "series": { "<item>": {"2021": 123.0, ...} }
    }
  }
}

- org_code는 data/institutions.json 의 기관명으로 매핑 (instCd)
- 매핑 실패 기관은 _meta.unmatched 에 기록
- 실행: .venv/Scripts/python scripts/build_metrics.py
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX_DIR = ROOT / "rawdata" / "xlsx"
OUT_DIR = ROOT / "data" / "metrics"
INSTITUTIONS = ROOT / "data" / "institutions.json"

YEAR_RE = re.compile(r"^(20\d{2})년?(?:\s*(남|여))?\s*$")
META_COLS = {"기관유형", "주무부처", "상위기관"}
CLASS_COLS = {"구분", "항목", "기금명", "연도", "고용형태", "상세내역"}

# 파일명 → (카테고리, 라벨, 단위, 시트별 항목 접두어 사용 여부)
FILE_SPEC: dict[str, dict] = {
    "임직원수현황.xlsx": {"category": "staff", "label": "임직원 수", "unit": "명"},
    "직원평균보수현황.xlsx": {"category": "salary", "label": "직원 평균보수", "unit": "천원"},
    "임원연봉.xlsx": {"category": "executive_pay", "label": "임원 연봉", "unit": "천원"},
    "신규채용현황.xlsx": {"category": "recruitment", "label": "신규채용·청년인턴", "unit": "명"},
    "수입지출현황.xlsx": {"category": "budget", "label": "수입·지출 현황", "unit": "백만원"},
    "복리후생비.xlsx": {"category": "welfare", "label": "복리후생비", "unit": "천원"},
    "일가정_양립_지원제도_운영현황.xlsx": {
        "category": "work_life", "label": "일·가정 양립 지원제도", "unit": "명(어린이집 운영비는 천원)",
    },
    "그밖의_복리후생제도_등의_운영현황.xlsx": {
        "category": "welfare_etc", "label": "그 밖의 복리후생제도", "unit": "천원",
    },
    "법인세정보.xlsx": {"category": "tax", "label": "법인세", "unit": "천원"},
    "기관장업무추진비.xlsx": {"category": "head_expense", "label": "기관장 업무추진비", "unit": "천원"},
}

HTML_XLS = "공기업_반기_재정현황_2026년_1분기.xls"


def load_name_to_code() -> dict[str, str]:
    data = json.loads(INSTITUTIONS.read_text(encoding="utf-8"))
    return {o["name"]: o["org_code"] for o in data["orgs"] if o.get("org_code")}


def to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v) if float(v).is_integer() else round(float(v), 2)
    s = str(v).replace(",", "").strip()
    if not s or s == "-":
        return None
    try:
        f = float(s)
        return int(f) if f.is_integer() else round(f, 2)
    except ValueError:
        return None


def find_header_row(sh) -> int | None:
    for row in sh.iter_rows(max_row=10):
        for cell in row:
            if cell.value == "기관명":
                return cell.row
    return None


def parse_sheet(sh, name_to_code, orgs: dict, unmatched: set, years: set, item_prefix: str):
    hdr_r = find_header_row(sh)
    if hdr_r is None:
        return 0
    headers = [c.value for c in sh[hdr_r]]

    org_col = None
    class_cols: list[int] = []
    year_cols: list[tuple[int, str, str | None]] = []  # (idx, year, gender)
    for i, h in enumerate(headers):
        if h is None:
            continue
        h = str(h).strip()
        if h == "기관명":
            org_col = i
        elif h in META_COLS:
            continue
        elif m := YEAR_RE.match(h):
            year_cols.append((i, m.group(1), m.group(2)))
        elif h in CLASS_COLS:
            class_cols.append(i)

    if org_col is None or not year_cols:
        return 0  # 특수 구조(직장어린이집 등)는 스킵

    n = 0
    for row in sh.iter_rows(min_row=hdr_r + 1, values_only=True):
        org_name = (str(row[org_col]).strip() if row[org_col] else "")
        if not org_name:
            continue
        code = name_to_code.get(org_name)
        if not code:
            unmatched.add(org_name)
            continue
        parts = [str(row[i]).strip() for i in class_cols if row[i] not in (None, "")]
        item = " | ".join(parts) if parts else "값"
        if item_prefix:
            item = f"{item_prefix} | {item}"

        org = orgs.setdefault(code, {"name": org_name, "series": {}})
        series = org["series"].setdefault(item, {})
        for i, year, gender in year_cols:
            val = to_num(row[i]) if i < len(row) else None
            if val is None:
                continue
            key = f"{year}-{gender}" if gender else year
            series[key] = val
            years.add(year)
        n += 1
    return n


def build_xlsx(path: Path, spec: dict, name_to_code) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    orgs: dict = {}
    unmatched: set = set()
    years: set = set()
    multi = len(wb.sheetnames) > 1
    skipped = []
    for sname in wb.sheetnames:
        prefix = re.sub(r"^\d+(-\d+)?\.\s*", "", sname).strip() if multi else ""
        n = parse_sheet(wb[sname], name_to_code, orgs, unmatched, years, prefix)
        if n == 0:
            skipped.append(sname)
    wb.close()

    # 빈 시리즈 제거
    for org in orgs.values():
        org["series"] = {k: v for k, v in org["series"].items() if v}
    orgs = {k: v for k, v in orgs.items() if v["series"]}

    caveats = [f"단위: {spec['unit']}", "원천: ALIO 항목별 공시 엑셀 (data.go.kr)"]
    if skipped:
        caveats.append(f"파싱 제외 시트(특수 구조): {skipped}")
    return {
        "_meta": {
            "category": spec["category"],
            "label": spec["label"],
            "unit": spec["unit"],
            "source_file": path.name,
            "years": sorted(years),
            "org_count": len(orgs),
            "built_at": datetime.now().isoformat(timespec="seconds"),
            "unmatched_orgs": sorted(unmatched),
            "caveats": caveats,
        },
        "orgs": orgs,
    }


def build_finance_html(path: Path, name_to_code) -> dict:
    """공기업 반기 재정현황 — HTML로 저장된 .xls 파싱."""
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "lxml")
    trs = soup.find("table").find_all("tr")

    # row2: 지표명 24개(colspan=5), row3: 연도 라벨 120개
    metric_names = [td.get_text(strip=True) for td in trs[2].find_all("td")]
    year_labels = [td.get_text(strip=True) for td in trs[3].find_all("td")]

    type_pat = re.compile(r"공기업|준정부|기타공공")
    orgs: dict = {}
    unmatched: set = set()
    years: set = set()

    for tr in trs[4:]:
        tds = tr.find_all("td")
        if len(tds) < len(year_labels) + 2:
            continue
        name = tds[0].get_text(strip=True)
        org_type = tds[1].get_text(strip=True)
        if not name or not type_pat.search(org_type):
            continue
        code = name_to_code.get(name)
        if not code:
            unmatched.add(name)
            continue
        values = [td.get_text(strip=True) for td in tds[2:2 + len(year_labels)]]
        series_map: dict[str, dict] = {}
        for j, label in enumerate(year_labels):
            metric = metric_names[j // 5] if j // 5 < len(metric_names) else f"지표{j}"
            m = re.search(r"(20\d{2})", label)
            if not m:
                continue
            year = m.group(1)
            val = to_num(values[j])
            if val is None:
                continue
            series_map.setdefault(metric, {})[year] = val
            years.add(year)
        if series_map:
            orgs[code] = {"name": name, "series": series_map}

    return {
        "_meta": {
            "category": "finance",
            "label": "요약 재무상태표·포괄손익계산서 (반기)",
            "unit": "백만원 (부채비율 등 비율 지표는 %)",
            "source_file": path.name,
            "years": sorted(years),
            "org_count": len(orgs),
            "built_at": datetime.now().isoformat(timespec="seconds"),
            "unmatched_orgs": sorted(unmatched),
            "caveats": [
                "단위: 백만원 (비율 지표는 %)",
                f"ALIO 반기 재정현황 공시 대상 {len(orgs)}개 기관 한정 — 과거 공기업 지정 이력 기관 포함 가능하므로 현행 공기업 수와 다를 수 있음",
                "전체 기관 재무는 추후 통합공시 PDF 파싱으로 보완 예정",
                "값은 각 연도 반기 기준",
            ],
        },
        "orgs": orgs,
    }


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    name_to_code = load_name_to_code()

    index = []
    for fname, spec in FILE_SPEC.items():
        path = XLSX_DIR / fname
        if not path.exists():
            print(f"skip (missing): {fname}")
            continue
        data = build_xlsx(path, spec, name_to_code)
        out = OUT_DIR / f"{spec['category']}.json"
        out.write_text(
            json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
        )
        m = data["_meta"]
        index.append({k: m[k] for k in ("category", "label", "unit", "years", "org_count")})
        print(f"{out.name}: orgs={m['org_count']} years={m['years']}"
              f" unmatched={len(m['unmatched_orgs'])} size={out.stat().st_size//1024}KB")

    html_path = XLSX_DIR / HTML_XLS
    if html_path.exists():
        data = build_finance_html(html_path, name_to_code)
        out = OUT_DIR / "finance.json"
        out.write_text(
            json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
        )
        m = data["_meta"]
        index.append({k: m[k] for k in ("category", "label", "unit", "years", "org_count")})
        print(f"{out.name}: orgs={m['org_count']} years={m['years']}"
              f" unmatched={len(m['unmatched_orgs'])}")

    (OUT_DIR / "_index.json").write_text(
        json.dumps(
            {"built_at": datetime.now().isoformat(timespec="seconds"), "categories": index},
            ensure_ascii=False, indent=1,
        ),
        encoding="utf-8",
    )
    print(f"index: {len(index)} categories")
    print("next: run `python -X utf8 scripts/promote_crawl_metrics.py` after `crawl_alio.py parse` to merge validated crawl values")


if __name__ == "__main__":
    main()
